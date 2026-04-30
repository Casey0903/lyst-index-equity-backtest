"""
Lyst Index → Stock Price Backtest v3
====================================
v3 changes from v2:
  1. Extended data: 2018 Q1 – latest (auto-detected from tracker CSV)
     - Reads xlsx (2019-2025) + v0.5 CSV (2018) + tracker CSV (2022+, fills 2026+)
  2. LVMH: only Louis Vuitton (was Loewe) — better revenue proxy
  3. Kering: only Gucci (was 5 brands) — dominant revenue contributor
  4. Percentile rank normalization for varying list sizes (10 vs 20)
  5. Train/test split: 2018-2021 vs 2022-latest (tuple compare, not lex)

2026-04-29: dynamic latest-quarter detection — reruns auto-extend as Lyst publishes new quarters.
"""

import csv
import math
import random
import statistics
from datetime import datetime, timedelta, date
from collections import defaultdict
import yfinance as yf

# ─── Config ───
# Dynamic price-fetch end (was hardcoded '2026-02-11', truncating Q4 2025 → Q1 2026 forward returns)
PRICE_FETCH_END = date.today().strftime("%Y-%m-%d")

from pathlib import Path
import os as _os

REPO_ROOT = Path(__file__).resolve().parent.parent
TRACKER_CSV = str(REPO_ROOT / "data" / "lyst-index-tracker.csv")
EXTENDED_CSV = str(REPO_ROOT / "data" / "lyst-historical-2018.csv")
# Optional third-party xlsx source — not included in the public repo.
# Set LYST_XLSX_PATH env var to use it locally; tracker CSV is used otherwise.
XLSX_PATH = _os.environ.get("LYST_XLSX_PATH", "")
OUTPUT_PATH = str(REPO_ROOT / "results" / "v3.1-results.md")

# ─── Quarter dates: auto-generated from Q1 2018 → latest quarter in tracker CSV ───
# Lyst publishes ~3-4 weeks after quarter end (stylized: month-after-quarter, day 25)
START_QUARTER = ("Y", 2018, 1)  # Q1 2018


def _parse_quarter_str(q_str):
    """'Q1 2026' -> (2026, 1)."""
    parts = q_str.strip().split()
    return int(parts[1]), int(parts[0][1:])


def _quarter_to_str(year, qtr):
    return f"Q{qtr} {year}"


def _quarter_end_date(year, qtr):
    end = {1: "03-31", 2: "06-30", 3: "09-30", 4: "12-31"}
    return f"{year}-{end[qtr]}"


def _publish_date(year, qtr):
    # Q1→04-25, Q2→07-25, Q3→10-25, Q4→next-year-01-25
    if qtr == 4:
        return f"{year + 1}-01-25"
    return f"{year}-{ {1: '04', 2: '07', 3: '10'}[qtr] }-25"


def _detect_latest_quarter():
    """Scan tracker CSV to find latest quarter present. Falls back to Q4 2025."""
    import os

    latest = (2025, 4)
    if os.path.exists(TRACKER_CSV):
        with open(TRACKER_CSV, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                q = row.get("Quarter", "").strip()
                if not q.startswith("Q"):
                    continue
                try:
                    yq = _parse_quarter_str(q)
                    if yq > latest:
                        latest = yq
                except (ValueError, IndexError):
                    continue
    return _quarter_to_str(*latest)


def _build_quarter_dicts(latest_q_str):
    """Generate QUARTER_END + PUBLISH_DATES from Q1 2018 to latest_q_str."""
    end_y, end_q = _parse_quarter_str(latest_q_str)
    qe, pd = {}, {}
    y, q = 2018, 1
    while (y, q) <= (end_y, end_q):
        q_str = _quarter_to_str(y, q)
        qe[q_str] = _quarter_end_date(y, q)
        pd[q_str] = _publish_date(y, q)
        y, q = (y + 1, 1) if q == 4 else (y, q + 1)
    return qe, pd


LATEST_QUARTER = _detect_latest_quarter()
QUARTER_END, PUBLISH_DATES = _build_quarter_dicts(LATEST_QUARTER)
print(
    f"[lyst-backtest] Latest quarter detected: {LATEST_QUARTER} ({len(QUARTER_END)} total)"
)

QUARTERS = list(QUARTER_END.keys())
CONSECUTIVE_PAIRS = [(QUARTERS[i], QUARTERS[i + 1]) for i in range(len(QUARTERS) - 1)]

# Train/test split: first half vs second half
# (Use tuple comparison via _parse_quarter_str — lex string compare on "QX YYYY" is broken)
TRAIN_QUARTERS = [
    q for q in QUARTERS if _parse_quarter_str(q) <= (2021, 4)
]  # 2018-2021
TEST_QUARTERS = [
    q for q in QUARTERS if _parse_quarter_str(q) >= (2022, 1)
]  # 2022-latest

# ─── Company config: ONLY LV for LVMH, ONLY Gucci for Kering ───
COMPANIES = {
    "Kering": {
        "ticker": "KER.PA",
        "currency": "EUR",
        "brands": {"Gucci": 1.0},  # v3: only Gucci (was 5 brands)
    },
    "Prada Group": {
        "ticker": "1913.HK",
        "currency": "HKD",
        "brands": {"Prada": 0.65, "Miu Miu": 0.35},
    },
    "LVMH": {
        "ticker": "MC.PA",
        "currency": "EUR",
        "brands": {"Louis Vuitton": 1.0},  # v3: only LV (was Loewe)
    },
    "Ralph Lauren": {
        "ticker": "RL",
        "currency": "USD",
        "brands": {"Ralph Lauren": 1.0},
    },
    "Burberry": {
        "ticker": "BRBY.L",
        "currency": "GBP",
        "brands": {"Burberry": 1.0},
    },
    "Moncler": {
        "ticker": "MONC.MI",
        "currency": "EUR",
        "brands": {
            "Moncler": 1.0,
            "Stone Island": 0.0,
        },  # Stone Island tracked but 0 weight until owned
    },
    "Tapestry": {
        "ticker": "TPR",
        "currency": "USD",
        "brands": {"Coach": 1.0},
    },
    "Nike": {
        "ticker": "NKE",
        "currency": "USD",
        "brands": {"Nike": 1.0},
    },
    # 2026-04-29: H&M Group removed — H&M (largest brand, ~75% revenue) never on Lyst (mass-market, not aspirational)
    # 2026-04-29: Capri removed — Michael Kors (largest brand, ~70%) never on Lyst (affordable luxury, not aspirational)
}

UNRANKED_RANK = 21  # Penalty for unranked brands


def _is_both_unranked(company_ranks, company, q1, q2):
    """True if company's composite rank == UNRANKED in both quarters (no real signal)."""
    r1 = (
        company_ranks[company][q1]["composite"]
        if q1 in company_ranks[company]
        else UNRANKED_RANK
    )
    r2 = (
        company_ranks[company][q2]["composite"]
        if q2 in company_ranks[company]
        else UNRANKED_RANK
    )
    return r1 >= UNRANKED_RANK and r2 >= UNRANKED_RANK


# ─── FX pairs for USD conversion ───
FX_TICKERS = {
    "EUR": "EURUSD=X",
    "GBP": "GBPUSD=X",
    "HKD": "HKDUSD=X",
    "SEK": "SEKUSD=X",
}


# ════════════════════════════════════════════
# Data Loading
# ════════════════════════════════════════════


def normalize_quarter(q_str):
    """Convert various quarter formats to 'Q1 2022' format."""
    q_str = q_str.strip()
    if q_str.startswith("Q"):
        return q_str  # Already normalized
    # Handle '2018Q1' format
    if "Q" in q_str:
        parts = q_str.split("Q")
        if len(parts) == 2:
            year, qtr = parts[0].strip(), parts[1].strip()
            return f"Q{qtr} {year}"
    # Handle '2023 Q4' format
    parts = q_str.split()
    if len(parts) == 2:
        return f"{parts[1]} {parts[0]}"
    return q_str


def parse_extended_data():
    """Parse the v0.5 extended CSV (2018 only)."""
    brand_data = defaultdict(dict)
    with open(EXTENDED_CSV, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            brand = row["brand"].strip().rstrip("?")
            quarter = normalize_quarter(row["quarter"])
            # Only use 2018 data from this source
            if quarter not in QUARTER_END or "2018" not in quarter:
                continue
            try:
                rank = int(row["rank"].strip())
            except (ValueError, KeyError):
                continue
            brand_data[brand][quarter] = rank
    return brand_data


def parse_tracker_csv():
    """Parse the main tracker CSV (Q1 2022 – latest) as fallback for missing xlsx."""
    brand_data = defaultdict(dict)
    import os

    if not os.path.exists(TRACKER_CSV):
        return brand_data
    with open(TRACKER_CSV, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            quarter = row["Quarter"].strip()
            if quarter not in QUARTER_END:
                continue
            brand = row["Brand"].strip()
            try:
                rank = int(row["Rank"].strip())
            except (ValueError, KeyError):
                continue
            # Normalize brand names
            brand_aliases = {
                "Alaïa": "Alaia",
                "Chloé": "Chloe",
                "Totême": "Toteme",
            }
            brand = brand_aliases.get(brand, brand)
            brand_data[brand][quarter] = rank
    return brand_data


def parse_xlsx_data():
    """Parse the xlsx file (1Q19–4Q25, pivoted format)."""
    import os

    if not os.path.exists(XLSX_PATH):
        print("  [WARN] xlsx not found, using tracker CSV as fallback")
        return parse_tracker_csv()
    import openpyxl

    brand_data = defaultdict(dict)
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["Brand Rankings"]

    # Read header row to get quarter columns
    header = [cell.value for cell in ws[1]]
    # header[0] = 'Rank', header[1] = '1Q19', header[2] = '2Q19', ...
    quarter_map = {}  # col_idx -> quarter string
    for col_idx, val in enumerate(header):
        if val is None or col_idx == 0:
            continue
        # Convert '1Q19' -> 'Q1 2019', '3Q20' -> 'Q3 2020'
        s = str(val).strip()
        if len(s) == 4 and "Q" in s:
            qtr_num = s[0]
            year_short = s[2:4]
            year = f"20{year_short}"
            q_str = f"Q{qtr_num} {year}"
            if q_str in QUARTER_END:
                quarter_map[col_idx] = q_str

    # Read data rows (row 2 onwards: rank 1-20)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        rank_val = row[0].value
        if rank_val is None:
            continue
        try:
            rank = int(rank_val)
        except (ValueError, TypeError):
            continue
        for col_idx, q_str in quarter_map.items():
            brand = row[col_idx].value
            if brand is None:
                continue
            brand = str(brand).strip()
            # Normalize brand names to match tracker CSV
            brand_aliases = {
                "Alaïa": "Alaia",
                "Chloé": "Chloe",
                "Totême": "Toteme",
                "SaintLaurent": "Saint Laurent",
                "Dolce Gabbana": "Dolce & Gabbana",
                "Dr. Martens": "Dr Martens",
            }
            brand = brand_aliases.get(brand, brand)
            brand_data[brand][q_str] = rank

    wb.close()
    return brand_data


def merge_brand_data(*sources):
    """Merge multiple brand data sources. Later sources override earlier ones."""
    merged = defaultdict(dict)
    for source in sources:
        for brand, quarters in source.items():
            for q, rank in quarters.items():
                merged[brand][q] = rank
    return merged


def compute_composite_ranks(brand_data, companies):
    """Revenue-weighted composite rank."""
    company_ranks = {}
    for company, cfg in companies.items():
        company_ranks[company] = {}
        for q in QUARTERS:
            weighted_rank = 0.0
            total_weight = 0.0
            best_rank = UNRANKED_RANK
            brands_ranked = 0
            for brand, weight in cfg["brands"].items():
                if weight == 0:  # Skip 0-weight brands
                    continue
                if brand in brand_data and q in brand_data[brand]:
                    r = brand_data[brand][q]
                    brands_ranked += 1
                else:
                    r = UNRANKED_RANK
                weighted_rank += r * weight
                total_weight += weight
                best_rank = min(best_rank, r)

            composite = (
                weighted_rank / total_weight if total_weight > 0 else UNRANKED_RANK
            )
            company_ranks[company][q] = {
                "composite": round(composite, 2),
                "best": best_rank,
                "brands_ranked": brands_ranked,
                "total_brands": sum(1 for w in cfg["brands"].values() if w > 0),
            }
    return company_ranks


def fetch_prices_and_fx():
    """Fetch stock prices + FX rates, return USD-denominated prices."""
    all_tickers = list(set(cfg["ticker"] for cfg in COMPANIES.values()))
    all_tickers.append("SPY")
    all_tickers.append("GLUX.PA")
    fx_tickers = list(FX_TICKERS.values())

    print(f"Fetching {len(all_tickers)} stocks + {len(fx_tickers)} FX pairs...")

    stock_data = {}
    for t in all_tickers:
        try:
            hist = yf.Ticker(t).history(
                start="2018-01-01", end=PRICE_FETCH_END, auto_adjust=True
            )
            if not hist.empty:
                stock_data[t] = hist
                print(f"  {t}: {len(hist)} days")
            else:
                print(f"  {t}: NO DATA")
        except Exception as e:
            print(f"  {t}: ERROR {e}")

    fx_data = {}
    for ccy, fx_tick in FX_TICKERS.items():
        try:
            hist = yf.Ticker(fx_tick).history(
                start="2018-01-01", end=PRICE_FETCH_END, auto_adjust=True
            )
            if not hist.empty:
                fx_data[ccy] = hist
                print(f"  FX {ccy}/USD: {len(hist)} days")
        except Exception as e:
            print(f"  FX {ccy}: ERROR {e}")

    return stock_data, fx_data


def get_price_after_date(hist, date_str, max_days=5):
    """Get closing price on the first trading day ON or AFTER a date."""
    target = datetime.strptime(date_str, "%Y-%m-%d")
    for delta in range(0, max_days + 1):
        check = (target + timedelta(days=delta)).strftime("%Y-%m-%d")
        matches = hist.loc[hist.index.strftime("%Y-%m-%d") == check]
        if not matches.empty:
            return float(matches["Close"].iloc[-1])
    return None


def _compute_benchmark_returns(stock_data, fx_data, ticker, currency="USD"):
    """Compute quarterly USD returns for a benchmark ticker."""
    returns = {}
    hist = stock_data.get(ticker)
    if hist is None:
        return returns
    for q1, q2 in CONSECUTIVE_PAIRS:
        p0 = get_price_after_date(hist, PUBLISH_DATES[q1])
        p1 = get_price_after_date(hist, PUBLISH_DATES[q2])
        if p0 is None or p1 is None:
            continue
        if currency != "USD" and currency in fx_data:
            fx_hist = fx_data[currency]
            fx0 = get_price_after_date(fx_hist, PUBLISH_DATES[q1])
            fx1 = get_price_after_date(fx_hist, PUBLISH_DATES[q2])
            if fx0 and fx1:
                p0 = p0 * fx0
                p1 = p1 * fx1
        returns[q1] = round((p1 / p0 - 1) * 100, 2)
    return returns


def compute_usd_returns(stock_data, fx_data, companies):
    """Compute forward returns in USD, only for consecutive quarter pairs."""
    raw_returns = {}

    spy_returns = _compute_benchmark_returns(stock_data, fx_data, "SPY", "USD")
    glux_returns = _compute_benchmark_returns(stock_data, fx_data, "GLUX.PA", "EUR")

    for company, cfg in companies.items():
        ticker = cfg["ticker"]
        ccy = cfg["currency"]
        hist = stock_data.get(ticker)
        if hist is None:
            continue
        raw_returns[company] = {}
        for q1, q2 in CONSECUTIVE_PAIRS:
            pub1 = PUBLISH_DATES[q1]
            pub2 = PUBLISH_DATES[q2]
            p0 = get_price_after_date(hist, pub1)
            p1 = get_price_after_date(hist, pub2)
            if p0 is None or p1 is None:
                continue
            if ccy != "USD" and ccy in fx_data:
                fx_hist = fx_data[ccy]
                fx0 = get_price_after_date(fx_hist, pub1)
                fx1 = get_price_after_date(fx_hist, pub2)
                if fx0 and fx1:
                    p0_usd = p0 * fx0
                    p1_usd = p1 * fx1
                else:
                    p0_usd = p0
                    p1_usd = p1
            else:
                p0_usd = p0
                p1_usd = p1
            raw_returns[company][q1] = round((p1_usd / p0_usd - 1) * 100, 2)

    ew_returns = {}
    for q1, _ in CONSECUTIVE_PAIRS:
        q_rets = [raw_returns[c][q1] for c in raw_returns if q1 in raw_returns[c]]
        if q_rets:
            ew_returns[q1] = round(statistics.mean(q_rets), 2)

    benchmarks = {
        "SPY": spy_returns,
        "GLUX": glux_returns,
        "EW": ew_returns,
    }

    all_excess = {}
    for bm_name, bm_rets in benchmarks.items():
        excess = {}
        for company in raw_returns:
            excess[company] = {}
            for q in raw_returns[company]:
                if q in bm_rets:
                    excess[company][q] = round(raw_returns[company][q] - bm_rets[q], 2)
        all_excess[bm_name] = excess

    return raw_returns, all_excess, benchmarks


# ════════════════════════════════════════════
# Statistical Helpers
# ════════════════════════════════════════════


def pearson_corr(xs, ys):
    if len(xs) < 3:
        return None
    n = len(xs)
    mx, my = sum(xs) / n, sum(ys) / n
    sx = (sum((x - mx) ** 2 for x in xs) / (n - 1)) ** 0.5
    sy = (sum((y - my) ** 2 for y in ys) / (n - 1)) ** 0.5
    if sx == 0 or sy == 0:
        return None
    cov = sum((xs[i] - mx) * (ys[i] - my) for i in range(n)) / (n - 1)
    return round(cov / (sx * sy), 4)


def t_stat_corr(r, n):
    if r is None or abs(r) >= 1.0 or n < 4:
        return None
    return r * math.sqrt(n - 2) / math.sqrt(1 - r**2)


def _norm_cdf(x):
    a1, a2, a3 = 0.4361836, -0.1201676, 0.9372980
    p = 0.33267
    t = 1 / (1 + p * abs(x))
    cdf = 1 - (a1 * t + a2 * t**2 + a3 * t**3) * math.exp(-x * x / 2) / math.sqrt(
        2 * math.pi
    )
    return cdf if x >= 0 else 1 - cdf


def _beta_incomplete(a, b, x):
    if x <= 0:
        return 0
    if x >= 1:
        return 1
    n_steps = 200
    dt = x / n_steps
    total = 0
    for i in range(n_steps):
        t = (i + 0.5) * dt
        if 0 < t < 1:
            total += t ** (a - 1) * (1 - t) ** (b - 1) * dt
    beta_val = math.gamma(a) * math.gamma(b) / math.gamma(a + b)
    return min(1.0, total / beta_val)


def p_value_t(t, df):
    if t is None:
        return None
    x = abs(t)
    if df > 30:
        p = 2 * (1 - _norm_cdf(x))
    else:
        a = df / (df + t * t)
        p = _beta_incomplete(df / 2, 0.5, a)
    return round(p, 4)


def t_test_mean(values):
    n = len(values)
    if n < 3:
        return None, None, None, n
    m = statistics.mean(values)
    s = statistics.stdev(values)
    if s == 0:
        return m, None, None, n
    t = m / (s / math.sqrt(n))
    p = p_value_t(t, n - 1)
    return round(m, 2), round(t, 3), p, n


def bootstrap_ci(values, n_boot=5000, ci=0.95):
    if len(values) < 3:
        return None, None
    means = []
    n = len(values)
    random.seed(42)
    for _ in range(n_boot):
        sample = [values[random.randint(0, n - 1)] for _ in range(n)]
        means.append(statistics.mean(sample))
    means.sort()
    lo = means[int(n_boot * (1 - ci) / 2)]
    hi = means[int(n_boot * (1 + ci) / 2) - 1]
    return round(lo, 2), round(hi, 2)


# ════════════════════════════════════════════
# Signals (use CONSECUTIVE_PAIRS to avoid gap)
# ════════════════════════════════════════════


def signal_s1_rank_improvement(company_ranks, returns, quarters=None):
    """S1: QoQ rank change → next-quarter excess return."""
    results = []
    for company in company_ranks:
        if company not in returns:
            continue
        for q1, q2 in CONSECUTIVE_PAIRS:
            if quarters and q1 not in quarters:
                continue
            idx = QUARTERS.index(q1)
            if idx == 0:
                continue
            prev_q = QUARTERS[idx - 1]
            if prev_q not in company_ranks[company] or q1 not in company_ranks[company]:
                continue
            if q1 not in returns[company]:
                continue
            # v3.1: skip if unranked in both quarters (no real signal)
            if _is_both_unranked(company_ranks, company, prev_q, q1):
                continue
            rank_change = (
                company_ranks[company][prev_q]["composite"]
                - company_ranks[company][q1]["composite"]
            )
            fwd_return = returns[company][q1]
            results.append(
                {
                    "company": company,
                    "quarter": q1,
                    "rank_change": round(rank_change, 2),
                    "fwd_return": fwd_return,
                    "direction_match": (rank_change > 0 and fwd_return > 0)
                    or (rank_change < 0 and fwd_return < 0),
                }
            )
    return results


def signal_s2_rank_level(company_ranks, returns, quarters=None):
    """S2: Absolute composite rank → next-quarter excess return."""
    results = []
    for company in company_ranks:
        if company not in returns:
            continue
        for q1, q2 in CONSECUTIVE_PAIRS:
            if quarters and q1 not in quarters:
                continue
            if q1 not in company_ranks[company] or q1 not in returns[company]:
                continue
            results.append(
                {
                    "company": company,
                    "quarter": q1,
                    "composite_rank": company_ranks[company][q1]["composite"],
                    "fwd_return": returns[company][q1],
                }
            )
    return results


def signal_s3_top10_entry(company_ranks, returns, quarters=None):
    """S3: Brand enters top 10 → next-quarter excess return."""
    results = []
    for company in company_ranks:
        if company not in returns:
            continue
        for q1, q2 in CONSECUTIVE_PAIRS:
            if quarters and q1 not in quarters:
                continue
            idx = QUARTERS.index(q1)
            if idx == 0:
                continue
            prev_q = QUARTERS[idx - 1]
            if prev_q not in company_ranks[company] or q1 not in company_ranks[company]:
                continue
            if q1 not in returns[company]:
                continue
            prev_best = company_ranks[company][prev_q]["best"]
            curr_best = company_ranks[company][q1]["best"]
            if prev_best > 10 and curr_best <= 10:
                results.append(
                    {
                        "company": company,
                        "quarter": q1,
                        "from_rank": prev_best,
                        "to_rank": curr_best,
                        "fwd_return": returns[company][q1],
                    }
                )
    return results


def signal_s3b_top10_exit(company_ranks, returns, quarters=None):
    """S3b: Brand exits top 10 → next-quarter excess return."""
    results = []
    for company in company_ranks:
        if company not in returns:
            continue
        for q1, q2 in CONSECUTIVE_PAIRS:
            if quarters and q1 not in quarters:
                continue
            idx = QUARTERS.index(q1)
            if idx == 0:
                continue
            prev_q = QUARTERS[idx - 1]
            if prev_q not in company_ranks[company] or q1 not in company_ranks[company]:
                continue
            if q1 not in returns[company]:
                continue
            prev_best = company_ranks[company][prev_q]["best"]
            curr_best = company_ranks[company][q1]["best"]
            if prev_best <= 10 and curr_best > 10:
                results.append(
                    {
                        "company": company,
                        "quarter": q1,
                        "from_rank": prev_best,
                        "to_rank": curr_best,
                        "fwd_return": returns[company][q1],
                    }
                )
    return results


def signal_s4_momentum(company_ranks, returns, quarters=None):
    """S4: 3+ consecutive quarters of rank improvement → next-quarter return."""
    results = []
    for company in company_ranks:
        if company not in returns:
            continue
        for i in range(3, len(QUARTERS)):
            q0 = QUARTERS[i]
            if quarters and q0 not in quarters:
                continue
            qs = [QUARTERS[i - 3], QUARTERS[i - 2], QUARTERS[i - 1], q0]
            if not all(q in company_ranks[company] for q in qs):
                continue
            if q0 not in returns.get(company, {}):
                continue

            ranks = [company_ranks[company][q]["composite"] for q in qs]
            # v3.1: skip if all 4 quarters are unranked
            if all(r >= UNRANKED_RANK for r in ranks):
                continue
            if ranks[0] > ranks[1] > ranks[2] > ranks[3]:
                results.append(
                    {
                        "company": company,
                        "quarter": q0,
                        "rank_trajectory": " → ".join(f"{r:.1f}" for r in ranks),
                        "total_improvement": round(ranks[0] - ranks[3], 1),
                        "fwd_return": returns[company][q0],
                    }
                )
    return results


def signal_s5_cross_sectional(company_ranks, returns, quarters=None):
    """S5: Cross-sectional long/short. Each quarter: long top 30% ΔRank, short bottom 30%."""
    portfolio_returns = []

    for q1, q2 in CONSECUTIVE_PAIRS:
        if quarters and q1 not in quarters:
            continue
        idx = QUARTERS.index(q1)
        if idx == 0:
            continue
        prev_q = QUARTERS[idx - 1]

        scored = []
        for company in company_ranks:
            if company not in returns:
                continue
            if q1 not in company_ranks[company] or prev_q not in company_ranks[company]:
                continue
            if q1 not in returns[company]:
                continue
            # v3.1: skip if unranked in both quarters (no real signal)
            if _is_both_unranked(company_ranks, company, prev_q, q1):
                continue
            rank_change = (
                company_ranks[company][prev_q]["composite"]
                - company_ranks[company][q1]["composite"]
            )
            scored.append(
                {
                    "company": company,
                    "rank_change": rank_change,
                    "fwd_return": returns[company][q1],
                }
            )

        if len(scored) < 4:
            continue

        scored.sort(key=lambda x: x["rank_change"], reverse=True)
        n = len(scored)
        top_n = max(1, n // 3)
        bot_n = max(1, n // 3)

        long_ret = statistics.mean([s["fwd_return"] for s in scored[:top_n]])
        short_ret = statistics.mean([s["fwd_return"] for s in scored[-bot_n:]])
        ls_ret = long_ret - short_ret

        portfolio_returns.append(
            {
                "quarter": q1,
                "long_ret": round(long_ret, 2),
                "short_ret": round(short_ret, 2),
                "ls_return": round(ls_ret, 2),
                "long_names": [s["company"] for s in scored[:top_n]],
                "short_names": [s["company"] for s in scored[-bot_n:]],
            }
        )

    return portfolio_returns


def signal_s6_interaction(company_ranks, returns, quarters=None):
    """S6: ΔRank × starting rank interaction."""
    results = []
    for company in company_ranks:
        if company not in returns:
            continue
        for q1, q2 in CONSECUTIVE_PAIRS:
            if quarters and q1 not in quarters:
                continue
            idx = QUARTERS.index(q1)
            if idx == 0:
                continue
            prev_q = QUARTERS[idx - 1]
            if prev_q not in company_ranks[company] or q1 not in company_ranks[company]:
                continue
            if q1 not in returns[company]:
                continue
            # v3.1: skip if unranked in both quarters (no real signal)
            if _is_both_unranked(company_ranks, company, prev_q, q1):
                continue
            prev_rank = company_ranks[company][prev_q]["composite"]
            curr_rank = company_ranks[company][q1]["composite"]
            rank_change = prev_rank - curr_rank
            results.append(
                {
                    "company": company,
                    "quarter": q1,
                    "starting_rank": round(prev_rank, 2),
                    "rank_change": round(rank_change, 2),
                    "interaction": round(rank_change * prev_rank, 2),
                    "fwd_return": returns[company][q1],
                    "zone": "bottom"
                    if prev_rank > 15
                    else ("mid" if prev_rank > 8 else "top"),
                }
            )
    return results


# ════════════════════════════════════════════
# Analysis & Reporting
# ════════════════════════════════════════════


def analyze_and_report(company_ranks, all_excess, raw_returns, benchmarks, brand_data):
    if benchmarks.get("GLUX"):
        primary_bm = "GLUX"
    else:
        primary_bm = "EW"
    excess_returns = all_excess[primary_bm]

    lines = []
    lines.append("---")
    lines.append("tags:\n  - backtest\n  - lyst-index\n  - consumer\n  - fashion")
    lines.append(f"date: {datetime.now().strftime('%Y-%m-%d')}")
    lines.append("version: v3.1")
    lines.append("---")
    lines.append("")
    lines.append("# Lyst Index → Stock Price Backtest v3.1")
    lines.append("")
    lines.append("> **v3.1 改进** (vs v3):")
    lines.append(
        "> - **双季 unranked 过滤**: 品牌连续两季不在 Lyst 榜时，排除该公司（ΔRank=0 为噪音非信号）"
    )
    lines.append(
        "> - 影响 S1, S4, S5, S6 信号（S3/S3b 不受影响，仅触发于 entry/exit 事件）"
    )
    lines.append(
        f"> - 数据/品牌/基准与 v3 完全一致: LVMH=LV, Kering=Gucci, {len(QUARTERS)} 季度（截止 {LATEST_QUARTER}）"
    )
    lines.append(f"> - 主基准: {primary_bm}")
    lines.append(f"> - Train/test 分割: 2018-2021 vs 2022-{LATEST_QUARTER.split()[1]}")
    lines.append(
        f"> - 总数据: {len(QUARTERS)} 个季度, {len(CONSECUTIVE_PAIRS)} 个连续季度对"
    )
    lines.append("")

    # ─── Data coverage summary ───
    lines.append("## 数据覆盖")
    lines.append("")
    lines.append("### v3 品牌变更影响")
    lines.append("")
    lines.append("| Company | v2 品牌 | v3 品牌 | 理由 |")
    lines.append("|---------|--------|--------|------|")
    lines.append(
        "| LVMH (MC.PA) | Loewe | **Louis Vuitton** | LV 占 LVMH Fashion & Leather ~60% 收入 |"
    )
    lines.append(
        "| Kering (KER.PA) | Gucci+SL+BV+Bal+Val | **Gucci only** | Gucci 占 Kering ~50% 收入 |"
    )
    lines.append("")

    # Brand rank coverage per quarter
    lines.append("### LV & Gucci 排名时间线")
    lines.append("")
    lines.append("| Quarter | Louis Vuitton | Gucci |")
    lines.append("|---------|--------------|-------|")
    for q in QUARTERS:
        lv_rank = brand_data.get("Louis Vuitton", {}).get(q, "—")
        gucci_rank = brand_data.get("Gucci", {}).get(q, "—")
        lines.append(f"| {q} | {lv_rank} | {gucci_rank} |")
    lines.append("")

    # ─── Benchmark comparison ───
    lines.append("## 基准对比")
    lines.append("")
    lines.append("| Quarter | SPY | GLUX (奢侈品) | EW (等权) | GLUX-SPY |")
    lines.append("|---------|-----|---------------|----------|----------|")
    for q1, q2 in CONSECUTIVE_PAIRS:
        spy_r = benchmarks["SPY"].get(q1)
        glux_r = benchmarks["GLUX"].get(q1)
        ew_r = benchmarks["EW"].get(q1)
        spy_s = f"{spy_r:+.2f}%" if spy_r is not None else "—"
        glux_s = f"{glux_r:+.2f}%" if glux_r is not None else "—"
        ew_s = f"{ew_r:+.2f}%" if ew_r is not None else "—"
        diff = (
            f"{glux_r - spy_r:+.2f}%"
            if (glux_r is not None and spy_r is not None)
            else "—"
        )
        lines.append(f"| {q1}→{q2} | {spy_s} | {glux_s} | {ew_s} | {diff} |")

    # Cumulative (handle gap by computing separately)
    for period_name, period_qs in [
        ("Train (2018-2021)", TRAIN_QUARTERS),
        (f"Test (2022-{LATEST_QUARTER.split()[1]})", TEST_QUARTERS),
        ("Overall", QUARTERS),
    ]:
        spy_cum = glux_cum = ew_cum = 1.0
        for q1, q2 in CONSECUTIVE_PAIRS:
            if q1 not in period_qs:
                continue
            if q1 in benchmarks["SPY"]:
                spy_cum *= 1 + benchmarks["SPY"][q1] / 100
            if q1 in benchmarks["GLUX"]:
                glux_cum *= 1 + benchmarks["GLUX"][q1] / 100
            if q1 in benchmarks["EW"]:
                ew_cum *= 1 + benchmarks["EW"][q1] / 100
        lines.append(
            f"| **{period_name}** | **{(spy_cum - 1) * 100:+.1f}%** | **{(glux_cum - 1) * 100:+.1f}%** | **{(ew_cum - 1) * 100:+.1f}%** | **{(glux_cum - spy_cum) * 100:+.1f}%** |"
        )
    lines.append("")

    # ─── Multi-benchmark robustness ───
    lines.append("### S1 信号在不同基准下的稳健性")
    lines.append("")
    lines.append(
        "| Benchmark | r (ΔRank vs Excess) | t-stat | p-value | Hit Rate | n |"
    )
    lines.append("|-----------|-------|--------|---------|----------|---|")
    for bm_name in ["SPY", "GLUX", "EW"]:
        bm_excess = all_excess[bm_name]
        s1_bm = signal_s1_rank_improvement(company_ranks, bm_excess)
        if s1_bm:
            xs = [r["rank_change"] for r in s1_bm]
            ys = [r["fwd_return"] for r in s1_bm]
            r_val = pearson_corr(xs, ys)
            t_v = t_stat_corr(r_val, len(s1_bm))
            p_v = p_value_t(t_v, len(s1_bm) - 2) if t_v else None
            hit = sum(1 for r in s1_bm if r["direction_match"]) / len(s1_bm) * 100
            t_str = f"{t_v:.2f}" if t_v else "—"
            p_str = f"{p_v:.3f}" if p_v else "—"
            bm_label = {
                "SPY": "SPY (大盘)",
                "GLUX": "GLUX (奢侈品指数)",
                "EW": "EW (等权)",
            }[bm_name]
            lines.append(
                f"| {bm_label} | {r_val} | {t_str} | {p_str} | {hit:.0f}% | {len(s1_bm)} |"
            )
    lines.append("")

    lines.append("### S3 信号在不同基准下的稳健性")
    lines.append("")
    lines.append("| Benchmark | Avg Excess | t-stat | p-value | % Positive | n |")
    lines.append("|-----------|-----------|--------|---------|-----------|---|")
    for bm_name in ["SPY", "GLUX", "EW"]:
        bm_excess = all_excess[bm_name]
        s3_bm = signal_s3_top10_entry(company_ranks, bm_excess)
        if s3_bm:
            rets = [r["fwd_return"] for r in s3_bm]
            m, t_v, p_v, _ = t_test_mean(rets)
            pct = sum(1 for r in rets if r > 0) / len(rets) * 100
            t_str = f"{t_v:.2f}" if t_v else "—"
            p_str = f"{p_v:.3f}" if p_v else "—"
            bm_label = {
                "SPY": "SPY (大盘)",
                "GLUX": "GLUX (奢侈品指数)",
                "EW": "EW (等权)",
            }[bm_name]
            lines.append(
                f"| {bm_label} | {m:+.2f}% | {t_str} | {p_str} | {pct:.0f}% | {len(s3_bm)} |"
            )
        else:
            bm_label = {
                "SPY": "SPY (大盘)",
                "GLUX": "GLUX (奢侈品指数)",
                "EW": "EW (等权)",
            }[bm_name]
            lines.append(f"| {bm_label} | — | — | — | — | 0 |")
    lines.append("")

    # ─── Full sample signals ───
    s1 = signal_s1_rank_improvement(company_ranks, excess_returns)
    s2 = signal_s2_rank_level(company_ranks, excess_returns)
    s3 = signal_s3_top10_entry(company_ranks, excess_returns)
    s3b = signal_s3b_top10_exit(company_ranks, excess_returns)
    s4 = signal_s4_momentum(company_ranks, excess_returns)
    s5 = signal_s5_cross_sectional(company_ranks, excess_returns)
    s6 = signal_s6_interaction(company_ranks, excess_returns)

    # Train/test (pre-gap vs post-gap)
    s1_train = signal_s1_rank_improvement(company_ranks, excess_returns, TRAIN_QUARTERS)
    s1_test = signal_s1_rank_improvement(company_ranks, excess_returns, TEST_QUARTERS)
    s5_train = signal_s5_cross_sectional(company_ranks, excess_returns, TRAIN_QUARTERS)
    s5_test = signal_s5_cross_sectional(company_ranks, excess_returns, TEST_QUARTERS)

    # ════════════ S1 ════════════
    lines.append("## S1: 排名改善 → 下季超额回报 (USD)")
    lines.append("")
    _analyze_s1(lines, s1, "全样本", company_ranks)
    lines.append("")
    lines.append("### Train/Test 分割")
    lines.append("")
    lines.append("**Train (2018 Q1 – 2021 Q4):**")
    _analyze_s1_brief(lines, s1_train)
    lines.append("")
    lines.append(f"**Test (2022 Q1 – {LATEST_QUARTER}):**")
    _analyze_s1_brief(lines, s1_test)
    lines.append("")

    # ════════════ S2 ════════════
    lines.append("## S2: 排名水平 → 下季超额回报 (Tercile)")
    lines.append("")
    if s2:
        s2_sorted = sorted(s2, key=lambda r: r["composite_rank"])
        n = len(s2_sorted)
        q_size = n // 3
        terciles = [
            ("Top (最佳排名)", s2_sorted[:q_size]),
            ("Mid (中等排名)", s2_sorted[q_size : 2 * q_size]),
            ("Bottom (最差排名)", s2_sorted[2 * q_size :]),
        ]
        lines.append(
            "| Tercile | Avg Excess | Median | t-stat | p-value | 95% CI | % Pos | n |"
        )
        lines.append(
            "|---------|-----------|--------|--------|---------|--------|-------|---|"
        )
        for name, group in terciles:
            rets = [r["fwd_return"] for r in group]
            m, t, p, n_obs = t_test_mean(rets)
            ci_lo, ci_hi = bootstrap_ci(rets)
            med = statistics.median(rets) if rets else 0
            pct_pos = sum(1 for r in rets if r > 0) / len(rets) * 100 if rets else 0
            t_str = f"{t:.2f}" if t else "—"
            p_str = f"{p:.3f}" if p else "—"
            ci_str = f"[{ci_lo}, {ci_hi}]" if ci_lo is not None else "—"
            m_val = m if m is not None else 0
            lines.append(
                f"| {name} | {m_val:+.2f}% | {med:+.2f}% | {t_str} | {p_str} | {ci_str} | {pct_pos:.0f}% | {len(rets)} |"
            )
    lines.append("")

    # ════════════ S3 ════════════
    lines.append("## S3: 进入 Top 10 → 下季超额回报")
    lines.append("")
    _analyze_event_signal(lines, s3, "进入 Top 10")
    lines.append("")

    # ════════════ S3b ════════════
    lines.append("## S3b: 跌出 Top 10 → 下季超额回报")
    lines.append("")
    _analyze_event_signal(lines, s3b, "跌出 Top 10")
    lines.append("")

    # ════════════ S4 ════════════
    lines.append("## S4: 连续3季上升 → 下季超额回报")
    lines.append("")
    if s4:
        rets = [r["fwd_return"] for r in s4]
        m, t, p, n_obs = t_test_mean(rets)
        ci_lo, ci_hi = bootstrap_ci(rets)
        pct_pos = sum(1 for r in rets if r > 0) / len(rets) * 100
        lines.append(f"- **事件数**: {len(s4)}")
        t_str = f"{t}" if t else "—"
        p_str = f"{p}" if p else "—"
        lines.append(f"- **平均超额回报**: {m:+.2f}%  (t={t_str}, p={p_str})")
        lines.append(f"- **95% CI**: [{ci_lo}, {ci_hi}]")
        lines.append(f"- **正超额率**: {pct_pos:.0f}%")
        lines.append("")
        lines.append("| Company | Quarter | Trajectory | Δ Rank | Excess Ret |")
        lines.append("|---------|---------|------------|--------|------------|")
        for r in s4:
            lines.append(
                f"| {r['company']} | {r['quarter']} | {r['rank_trajectory']} | {r['total_improvement']:+.1f} | {r['fwd_return']:+.2f}% |"
            )
    else:
        lines.append("无连续3季上升事件。")
    lines.append("")

    # ════════════ S5 ════════════
    lines.append("## S5: 截面多空组合")
    lines.append("")
    if s5:
        ls_rets = [r["ls_return"] for r in s5]
        m, t, p, n_obs = t_test_mean(ls_rets)
        ci_lo, ci_hi = bootstrap_ci(ls_rets)

        # Cumulative by period
        for period_name, period_qs in [
            ("Train (2018-2021)", TRAIN_QUARTERS),
            (f"Test (2022-{LATEST_QUARTER.split()[1]})", TEST_QUARTERS),
            ("Overall", QUARTERS),
        ]:
            cum_ret = 1.0
            period_rets = [r["ls_return"] for r in s5 if r["quarter"] in period_qs]
            for lr in period_rets:
                cum_ret *= 1 + lr / 100
            lines.append(
                f"**{period_name}** 累计 L/S: {(cum_ret - 1) * 100:+.1f}% ({len(period_rets)} quarters)"
            )

        lines.append("")
        t_str = f"{t}" if t else "—"
        p_str = f"{p}" if p else "—"
        lines.append(f"- **平均季度多空回报**: {m:+.2f}% (t={t_str}, p={p_str})")
        lines.append(f"- **95% CI**: [{ci_lo}, {ci_hi}]")
        lines.append("")
        lines.append("| Quarter | Long | Short | L/S Return |")
        lines.append("|---------|------|-------|------------|")
        for r in s5:
            long_str = ", ".join(r["long_names"][:3])
            short_str = ", ".join(r["short_names"][:3])
            lines.append(
                f"| {r['quarter']} | {long_str} | {short_str} | {r['ls_return']:+.2f}% |"
            )

        lines.append("")
        lines.append("### Train/Test 分割")
        if s5_train:
            ls_train = [r["ls_return"] for r in s5_train]
            m_tr, t_tr, p_tr, _ = t_test_mean(ls_train)
            t_str = f"{t_tr}" if t_tr else "—"
            p_str = f"{p_tr}" if p_tr else "—"
            lines.append(
                f"- **Train (2018-2021)**: avg={m_tr:+.2f}%, t={t_str}, p={p_str}, n={len(ls_train)}"
            )
        if s5_test:
            ls_test = [r["ls_return"] for r in s5_test]
            m_te, t_te, p_te, _ = t_test_mean(ls_test)
            t_str = f"{t_te}" if t_te else "—"
            p_str = f"{p_te}" if p_te else "—"
            lines.append(
                f"- **Test (2022-{LATEST_QUARTER.split()[1]})**: avg={m_te:+.2f}%, t={t_str}, p={p_str}, n={len(ls_test)}"
            )
    lines.append("")

    # ════════════ S6 ════════════
    lines.append("## S6: ΔRank × 起始排名交互")
    lines.append("")
    if s6:
        for zone in ["top", "mid", "bottom"]:
            zone_data = [r for r in s6 if r["zone"] == zone]
            if len(zone_data) < 3:
                continue
            xs = [r["rank_change"] for r in zone_data]
            ys = [r["fwd_return"] for r in zone_data]
            r_val = pearson_corr(xs, ys)
            t_v = t_stat_corr(r_val, len(zone_data))
            p_v = p_value_t(t_v, len(zone_data) - 2) if t_v else None
            zone_label = {"top": "Top 8", "mid": "Mid 9-15", "bottom": "Bottom 15+"}[
                zone
            ]
            r_str = f"{r_val:+.3f}" if r_val else "N/A"
            t_str = f"{t_v:.2f}" if t_v else "—"
            p_str = f"{p_v:.3f}" if p_v else "—"
            lines.append(
                f"**{zone_label}** (n={len(zone_data)}): r={r_str}, t={t_str}, p={p_str}"
            )

        xs = [r["interaction"] for r in s6]
        ys = [r["fwd_return"] for r in s6]
        r_int = pearson_corr(xs, ys)
        t_int = t_stat_corr(r_int, len(s6))
        p_int = p_value_t(t_int, len(s6) - 2) if t_int else None
        lines.append("")
        r_str = f"{r_int:+.3f}" if r_int else "N/A"
        t_str = f"{t_int:.2f}" if t_int else "—"
        p_str = f"{p_int:.3f}" if p_int else "—"
        lines.append(
            f"**交互项 (ΔRank × StartRank) → Excess Return**: r={r_str}, t={t_str}, p={p_str}, n={len(s6)}"
        )
    lines.append("")

    # ════════════ Summary table ════════════
    lines.append("## 信号汇总")
    lines.append("")
    lines.append(
        "| Signal | Description | Key Metric | t-stat | p-value | Significant? |"
    )
    lines.append(
        "|--------|-------------|-----------|--------|---------|-------------|"
    )

    if s1:
        xs = [r["rank_change"] for r in s1]
        ys = [r["fwd_return"] for r in s1]
        r_val = pearson_corr(xs, ys)
        t_v = t_stat_corr(r_val, len(s1))
        p_v = p_value_t(t_v, len(s1) - 2) if t_v else None
        sig = "✓" if p_v and p_v < 0.05 else "✗"
        t_str = f"{t_v:.2f}" if t_v else "—"
        p_str = f"{p_v:.3f}" if p_v else "—"
        lines.append(
            f"| S1 | ΔRank → Excess Ret | r={r_val} | {t_str} | {p_str} | {sig} |"
        )

    if s3:
        rets = [r["fwd_return"] for r in s3]
        m, t_v, p_v, _ = t_test_mean(rets)
        sig = "✓" if p_v and p_v < 0.05 else "✗"
        t_str = f"{t_v:.2f}" if t_v else "—"
        p_str = f"{p_v:.3f}" if p_v else "—"
        lines.append(
            f"| S3 | Top-10 Entry | avg={m:+.2f}% | {t_str} | {p_str} | {sig} |"
        )

    if s3b:
        rets = [r["fwd_return"] for r in s3b]
        m, t_v, p_v, _ = t_test_mean(rets)
        sig = "✓" if p_v and p_v < 0.05 else "✗"
        t_str = f"{t_v:.2f}" if t_v else "—"
        p_str = f"{p_v:.3f}" if p_v else "—"
        lines.append(
            f"| S3b | Top-10 Exit | avg={m:+.2f}% | {t_str} | {p_str} | {sig} |"
        )

    if s5:
        ls_rets = [r["ls_return"] for r in s5]
        m, t_v, p_v, _ = t_test_mean(ls_rets)
        sig = "✓" if p_v and p_v < 0.05 else "✗"
        t_str = f"{t_v:.2f}" if t_v else "—"
        p_str = f"{p_v:.3f}" if p_v else "—"
        lines.append(
            f"| S5 | Cross-sectional L/S | avg={m:+.2f}% | {t_str} | {p_str} | {sig} |"
        )

    lines.append("")

    # ════════════ Conclusion ════════════
    lines.append("## 结论与投资启示")
    lines.append("")
    lines.append("*v3.1 自动生成 — 所有回报为 USD 超额回报。*")
    lines.append(
        "*v3.1 关键变更: 双季 unranked 过滤 — 品牌连续两季不在榜时排除该公司，消除 ΔRank=0 噪音。*"
    )
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append(
        "数据: [[lyst-index-tracker]] | v3: [[lyst-backtest-v3-results]] | v2: [[lyst-backtest-v2-results]]"
    )

    return "\n".join(lines)


def _analyze_s1(lines, s1, label, company_ranks):
    if not s1:
        lines.append("无数据。")
        return
    xs = [r["rank_change"] for r in s1]
    ys = [r["fwd_return"] for r in s1]
    corr = pearson_corr(xs, ys)
    t = t_stat_corr(corr, len(s1))
    p = p_value_t(t, len(s1) - 2) if t else None
    hit_rate = sum(1 for r in s1 if r["direction_match"]) / len(s1) * 100

    improvers = [r for r in s1 if r["rank_change"] > 1]
    decliners = [r for r in s1 if r["rank_change"] < -1]
    avg_imp = statistics.mean([r["fwd_return"] for r in improvers]) if improvers else 0
    avg_dec = statistics.mean([r["fwd_return"] for r in decliners]) if decliners else 0

    t_str = f"{t:.2f}" if t else "—"
    p_str = f"{p:.3f}" if p else "—"

    lines.append(f"**{label}** (n={len(s1)})")
    lines.append("")
    lines.append(f"- **相关系数**: r={corr} (t={t_str}, p={p_str})")
    lines.append(f"- **方向命中率**: {hit_rate:.1f}%")
    lines.append(
        f"- **改善组**: {avg_imp:+.2f}% (n={len(improvers)}) vs **下跌组**: {avg_dec:+.2f}% (n={len(decliners)})"
    )
    lines.append(f"- **多空价差**: {avg_imp - avg_dec:+.2f}%")
    lines.append("")

    lines.append(
        "| Company | Ticker | n | Corr (r) | t-stat | p-value | Hit% | Avg Excess |"
    )
    lines.append(
        "|---------|--------|---|----------|--------|---------|------|-----------|"
    )
    for company in sorted(COMPANIES.keys()):
        co_data = [r for r in s1 if r["company"] == company]
        if len(co_data) < 3:
            continue
        cxs = [r["rank_change"] for r in co_data]
        cys = [r["fwd_return"] for r in co_data]
        r_val = pearson_corr(cxs, cys)
        t_co = t_stat_corr(r_val, len(co_data))
        p_co = p_value_t(t_co, len(co_data) - 2) if t_co else None
        hit = sum(1 for r in co_data if r["direction_match"]) / len(co_data) * 100
        avg_ret = statistics.mean(cys)
        ticker = COMPANIES[company]["ticker"]
        r_str = f"{r_val:+.3f}" if r_val else "N/A"
        t_str = f"{t_co:.2f}" if t_co else "—"
        p_str = f"{p_co:.3f}" if p_co else "—"
        lines.append(
            f"| {company} | {ticker} | {len(co_data)} | {r_str} | {t_str} | {p_str} | {hit:.0f}% | {avg_ret:+.2f}% |"
        )


def _analyze_s1_brief(lines, s1_sub):
    if not s1_sub:
        lines.append("  无数据。")
        return
    xs = [r["rank_change"] for r in s1_sub]
    ys = [r["fwd_return"] for r in s1_sub]
    corr = pearson_corr(xs, ys)
    t = t_stat_corr(corr, len(s1_sub))
    p = p_value_t(t, len(s1_sub) - 2) if t else None
    hit = sum(1 for r in s1_sub if r["direction_match"]) / len(s1_sub) * 100
    t_str = f"{t:.2f}" if t else "—"
    p_str = f"{p:.3f}" if p else "—"
    lines.append(f"  r={corr}, t={t_str}, p={p_str}, hit={hit:.0f}%, n={len(s1_sub)}")


def _analyze_event_signal(lines, events, label):
    if not events:
        lines.append(f"无{label}事件。")
        return
    rets = [r["fwd_return"] for r in events]
    m, t, p, n = t_test_mean(rets)
    ci_lo, ci_hi = bootstrap_ci(rets)
    pct_pos = sum(1 for r in rets if r > 0) / len(rets) * 100

    t_str = f"{t:.2f}" if t else "—"
    p_str = f"{p:.3f}" if p else "—"
    ci_str = f"[{ci_lo}, {ci_hi}]" if ci_lo is not None else "—"

    lines.append(f"- **事件数**: {len(events)}")
    lines.append(f"- **平均超额回报**: {m:+.2f}% (t={t_str}, p={p_str})")
    lines.append(f"- **95% CI**: {ci_str}")
    lines.append(f"- **正超额率**: {pct_pos:.0f}%")
    lines.append("")
    lines.append("| Company | Quarter | Rank Move | Excess Return |")
    lines.append("|---------|---------|-----------|---------------|")
    for r in events:
        lines.append(
            f"| {r['company']} | {r['quarter']} | {r['from_rank']}→{r['to_rank']} | {r['fwd_return']:+.2f}% |"
        )


# ════════════════════════════════════════════
# Main
# ════════════════════════════════════════════


def main():
    print("=" * 60)
    print("LYST INDEX → STOCK PRICE BACKTEST v3")
    print("  LVMH = Louis Vuitton only")
    print("  Kering = Gucci only")
    print(
        f"  Data: 2018 Q1 – {LATEST_QUARTER} ({len(QUARTERS)} quarters, auto-detected)"
    )
    print("=" * 60)

    # Parse all data sources
    extended_data = parse_extended_data()
    print(f"\nParsed {len(extended_data)} brands from extended CSV (2018)")
    tracker_data = parse_tracker_csv()
    print(f"Parsed {len(tracker_data)} brands from tracker CSV (2022-latest)")
    xlsx_data = parse_xlsx_data()
    print(f"Parsed {len(xlsx_data)} brands from xlsx (2019-2025)")

    # Merge order (later overrides earlier):
    #   extended_data: 2018 only
    #   tracker_data:  2022-latest (covers 2026+ that xlsx lacks)
    #   xlsx_data:     1Q19-4Q25 (authoritative for that range)
    brand_data = merge_brand_data(extended_data, tracker_data, xlsx_data)
    print(f"Merged: {len(brand_data)} total brands")

    # Check key brands
    for brand in ["Louis Vuitton", "Gucci", "Prada", "Miu Miu"]:
        quarters_with_data = sorted([q for q in brand_data.get(brand, {})])
        if quarters_with_data:
            print(
                f"  {brand}: {len(quarters_with_data)} quarters ({quarters_with_data[0]} – {quarters_with_data[-1]})"
            )
        else:
            print(f"  {brand}: NO DATA")

    print(f"\nQuarters: {len(QUARTERS)}")
    print(f"Consecutive pairs: {len(CONSECUTIVE_PAIRS)}")
    print(f"Train (2018-2021): {len(TRAIN_QUARTERS)} quarters")
    print(f"Test (2022-{LATEST_QUARTER.split()[1]}): {len(TEST_QUARTERS)} quarters")

    # Composite ranks
    company_ranks = compute_composite_ranks(brand_data, COMPANIES)
    print("\nComposite ranks computed")

    # Fetch prices
    stock_data, fx_data = fetch_prices_and_fx()

    # USD returns
    raw_returns, all_excess, benchmarks = compute_usd_returns(
        stock_data, fx_data, COMPANIES
    )
    print(f"\nComputed USD returns for {len(raw_returns)} companies")

    # Print benchmark summary
    for bm_name, bm_rets in benchmarks.items():
        print(f"\n{bm_name} quarterly returns:")
        cum = 1.0
        for q1, q2 in CONSECUTIVE_PAIRS:
            if q1 in bm_rets:
                print(f"  {q1}: {bm_rets[q1]:+.2f}%")
                cum *= 1 + bm_rets[q1] / 100
        print(f"  Cumulative: {(cum - 1) * 100:+.1f}%")

    # Generate report
    report = analyze_and_report(
        company_ranks, all_excess, raw_returns, benchmarks, brand_data
    )

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write(report)

    print(f"\n{'=' * 60}")
    print(f"Report saved to: {OUTPUT_PATH}")
    print("=" * 60)


if __name__ == "__main__":
    main()
